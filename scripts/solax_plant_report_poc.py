"""Proof-of-concept parser for SolaX Plant Report XLSX exports.

The parser treats report energy columns as cumulative daily readings and
converts them into interval kWh values by differencing adjacent readings
within each local day. It never writes to the raw input directory.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
from collections.abc import Iterable
from dataclasses import dataclass
from datetime import datetime, time, timedelta
from pathlib import Path

from openpyxl import load_workbook

DEFAULT_INPUT_DIR = Path("data/raw/solax")
DEFAULT_OUTPUT_DIR = Path("data/processed")
EXPECTED_STEP = timedelta(minutes=5)
TIMESTAMP_HEADER = "Update time"
HEADER_ROW = 2
FIRST_DATA_ROW = 3

ENERGY_COLUMNS = {
    "Daily PV Yield(kWh)": "daily_pv_yield_kwh",
    "Daily inverter output (kWh)": "daily_inverter_output_kwh",
    "Daily exported energy(kWh)": "daily_exported_energy_kwh",
    "Daily consumed(kWh)": "daily_consumed_kwh",
    "Daily imported energy(kWh)": "daily_imported_energy_kwh",
}


@dataclass(frozen=True)
class PlantReport:
    source_file_id: str
    source_file_hash: str
    sheet_name: str
    metadata: dict[str, str]
    headers: list[str]
    rows: list[dict[str, object]]


def redact(value: object) -> str:
    """Return a conservative redacted string for metadata/report text."""
    text = "" if value is None else str(value)
    text = re.sub(r"[\w.+-]+@[\w.-]+", "[REDACTED_EMAIL]", text)
    text = re.sub(r"\b[A-Z0-9]{8,}\b", "[REDACTED_ID]", text, flags=re.IGNORECASE)
    return text


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def parse_timestamp(value: object) -> datetime:
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    if value is None:
        raise ValueError("missing timestamp")
    text = str(value).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%Y-%m-%d %H:%M"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            pass
    raise ValueError(f"unsupported timestamp format: {text!r}")


def parse_number(value: object) -> float | None:
    if value is None or value == "":
        return None
    return float(value)


def find_report_files(input_dir: Path) -> list[Path]:
    if not input_dir.exists():
        return []
    return sorted(path for path in input_dir.glob("*.xlsx") if not path.name.startswith("~$"))


def load_report(path: Path, source_file_id: str) -> PlantReport:
    workbook = load_workbook(path, read_only=False, data_only=True)
    try:
        sheet = workbook.worksheets[0]
        headers = [
            "" if cell.value is None else str(cell.value).strip() for cell in sheet[HEADER_ROW]
        ]
        if TIMESTAMP_HEADER not in headers:
            raise ValueError(f"missing required header {TIMESTAMP_HEADER!r}")

        missing = [header for header in ENERGY_COLUMNS if header not in headers]
        if missing:
            raise ValueError(f"missing required energy headers: {missing}")

        metadata = {
            "source_file_id": source_file_id,
            "sheet_name": sheet.title,
            "raw_metadata_redacted": redact(sheet.cell(1, 1).value),
            "raw_metadata_sha256": hashlib.sha256(
                str(sheet.cell(1, 1).value or "").encode("utf-8")
            ).hexdigest(),
            "row_count": str(sheet.max_row),
            "column_count": str(sheet.max_column),
        }

        rows: list[dict[str, object]] = []
        for row_number in range(FIRST_DATA_ROW, sheet.max_row + 1):
            values = {
                headers[col_index - 1]: sheet.cell(row_number, col_index).value
                for col_index in range(1, len(headers) + 1)
                if headers[col_index - 1]
            }
            if not values.get(TIMESTAMP_HEADER):
                continue
            parsed = {
                "source_file_id": source_file_id,
                "row_number": row_number,
                "timestamp": parse_timestamp(values[TIMESTAMP_HEADER]),
            }
            for original, normalized in ENERGY_COLUMNS.items():
                parsed[normalized] = parse_number(values.get(original))
            rows.append(parsed)

        return PlantReport(
            source_file_id=source_file_id,
            source_file_hash=file_sha256(path),
            sheet_name=sheet.title,
            metadata=metadata,
            headers=headers,
            rows=rows,
        )
    finally:
        workbook.close()


def event(
    source_file_id: str,
    event_type: str,
    timestamp: datetime | None,
    field: str | None,
    message: str,
) -> dict[str, str]:
    return {
        "source_file_id": source_file_id,
        "event_type": event_type,
        "timestamp": "" if timestamp is None else timestamp.isoformat(sep=" "),
        "field": "" if field is None else field,
        "message": message,
    }


def detect_missing_timestamps(
    source_file_id: str, timestamps: list[datetime]
) -> list[dict[str, str]]:
    events: list[dict[str, str]] = []
    seen = sorted(set(timestamps))
    for previous, current in zip(seen, seen[1:], strict=False):
        if previous.date() != current.date():
            continue
        expected = previous + EXPECTED_STEP
        while expected < current:
            events.append(
                event(
                    source_file_id,
                    "missing_timestamp",
                    expected,
                    None,
                    "Expected five-minute timestamp is absent.",
                )
            )
            expected += EXPECTED_STEP
    return events


def dedupe_rows(report: PlantReport) -> tuple[list[dict[str, object]], list[dict[str, str]]]:
    events: list[dict[str, str]] = []
    deduped: dict[datetime, dict[str, object]] = {}
    for row in report.rows:
        timestamp = row["timestamp"]
        assert isinstance(timestamp, datetime)
        if timestamp in deduped:
            events.append(
                event(
                    report.source_file_id,
                    "duplicate_timestamp",
                    timestamp,
                    None,
                    "Duplicate timestamp skipped during interval conversion.",
                )
            )
            continue
        deduped[timestamp] = row
    return [deduped[key] for key in sorted(deduped)], events


def build_intervals(report: PlantReport) -> tuple[list[dict[str, object]], list[dict[str, str]]]:
    rows, events = dedupe_rows(report)
    timestamps = [row["timestamp"] for row in rows if isinstance(row["timestamp"], datetime)]
    events.extend(detect_missing_timestamps(report.source_file_id, timestamps))

    intervals: list[dict[str, object]] = []
    previous_by_day: dict[datetime.date, dict[str, object]] = {}
    previous_row: dict[str, object] | None = None

    for row in rows:
        timestamp = row["timestamp"]
        assert isinstance(timestamp, datetime)
        day = timestamp.date()

        if previous_row is not None:
            previous_timestamp = previous_row["timestamp"]
            assert isinstance(previous_timestamp, datetime)
            if previous_timestamp.date() != day:
                for field in ENERGY_COLUMNS.values():
                    previous_value = previous_row.get(field)
                    current_value = row.get(field)
                    if (
                        isinstance(previous_value, int | float)
                        and isinstance(current_value, int | float)
                        and current_value <= previous_value
                    ):
                        events.append(
                            event(
                                report.source_file_id,
                                "midnight_reset",
                                timestamp,
                                field,
                                "Cumulative daily reading reset at day boundary.",
                            )
                        )

        previous_same_day = previous_by_day.get(day)
        if previous_same_day is None:
            day_start = datetime.combine(day, time.min)
            if timestamp > day_start:
                for field in ENERGY_COLUMNS.values():
                    current_value = row.get(field)
                    if isinstance(current_value, int | float) and current_value != 0:
                        intervals.append(
                            interval_row(
                                report.source_file_id, day_start, timestamp, field, current_value
                            )
                        )
        else:
            previous_timestamp = previous_same_day["timestamp"]
            assert isinstance(previous_timestamp, datetime)
            for field in ENERGY_COLUMNS.values():
                previous_value = previous_same_day.get(field)
                current_value = row.get(field)
                if not isinstance(previous_value, int | float) or not isinstance(
                    current_value, int | float
                ):
                    continue
                diff = current_value - previous_value
                if diff < -1e-9:
                    events.append(
                        event(
                            report.source_file_id,
                            "negative_difference",
                            timestamp,
                            field,
                            "Cumulative daily reading decreased within the same day.",
                        )
                    )
                    continue
                intervals.append(
                    interval_row(report.source_file_id, previous_timestamp, timestamp, field, diff)
                )

        previous_by_day[day] = row
        previous_row = row

    return intervals, events


def interval_row(
    source_file_id: str,
    interval_start: datetime,
    interval_end: datetime,
    field: str,
    interval_kwh: float,
) -> dict[str, object]:
    return {
        "source_file_id": source_file_id,
        "interval_start": interval_start,
        "interval_end": interval_end,
        "date": interval_end.date().isoformat(),
        "field": field,
        "interval_kwh": round(float(interval_kwh), 6),
    }


def validate_daily_totals(
    report: PlantReport, intervals: Iterable[dict[str, object]]
) -> list[dict[str, object]]:
    interval_totals: dict[tuple[str, str], float] = {}
    for interval in intervals:
        key = (str(interval["date"]), str(interval["field"]))
        interval_totals[key] = interval_totals.get(key, 0.0) + float(interval["interval_kwh"])

    final_by_day_field: dict[tuple[str, str], float] = {}
    for row in report.rows:
        timestamp = row["timestamp"]
        if not isinstance(timestamp, datetime):
            continue
        for field in ENERGY_COLUMNS.values():
            value = row.get(field)
            if isinstance(value, int | float):
                final_by_day_field[(timestamp.date().isoformat(), field)] = float(value)

    summaries: list[dict[str, object]] = []
    for key in sorted(final_by_day_field):
        final_value = final_by_day_field[key]
        reconstructed = interval_totals.get(key, 0.0)
        difference = reconstructed - final_value
        summaries.append(
            {
                "source_file_id": report.source_file_id,
                "date": key[0],
                "field": key[1],
                "final_cumulative_kwh": round(final_value, 6),
                "reconstructed_interval_kwh": round(reconstructed, 6),
                "difference_kwh": round(difference, 6),
                "matches_final_cumulative": abs(difference) <= 0.001,
            }
        )
    return summaries


def write_csv(path: Path, rows: list[dict[str, object]], fieldnames: list[str]) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(
                {
                    key: value.isoformat(sep=" ") if isinstance(value, datetime) else value
                    for key, value in row.items()
                }
            )


def process_reports(input_dir: Path, output_dir: Path) -> dict[str, object]:
    output_dir.mkdir(parents=True, exist_ok=True)
    files = find_report_files(input_dir)

    metadata_rows: list[dict[str, object]] = []
    cumulative_rows: list[dict[str, object]] = []
    interval_rows: list[dict[str, object]] = []
    event_rows: list[dict[str, str]] = []
    validation_rows: list[dict[str, object]] = []

    for index, path in enumerate(files, start=1):
        source_file_id = f"solax_report_{index:03d}"
        try:
            report = load_report(path, source_file_id)
        except Exception as exc:
            event_rows.append(
                event(
                    source_file_id,
                    "file_error",
                    None,
                    None,
                    f"{type(exc).__name__}: unable to read workbook.",
                )
            )
            continue

        metadata_rows.append(
            {
                **report.metadata,
                "source_file_hash": report.source_file_hash,
                "headers": "|".join(report.headers),
            }
        )
        cumulative_rows.extend(report.rows)
        intervals, events = build_intervals(report)
        interval_rows.extend(intervals)
        event_rows.extend(events)
        validation_rows.extend(validate_daily_totals(report, intervals))

    write_csv(
        output_dir / "solax_report_metadata.csv",
        metadata_rows,
        [
            "source_file_id",
            "sheet_name",
            "raw_metadata_redacted",
            "raw_metadata_sha256",
            "row_count",
            "column_count",
            "source_file_hash",
            "headers",
        ],
    )
    write_csv(
        output_dir / "solax_cumulative_rows.csv",
        cumulative_rows,
        ["source_file_id", "row_number", "timestamp", *ENERGY_COLUMNS.values()],
    )
    write_csv(
        output_dir / "solax_interval_energy.csv",
        interval_rows,
        ["source_file_id", "interval_start", "interval_end", "date", "field", "interval_kwh"],
    )
    write_csv(
        output_dir / "solax_validation_events.csv",
        event_rows,
        ["source_file_id", "event_type", "timestamp", "field", "message"],
    )
    write_csv(
        output_dir / "solax_daily_validation_summary.csv",
        validation_rows,
        [
            "source_file_id",
            "date",
            "field",
            "final_cumulative_kwh",
            "reconstructed_interval_kwh",
            "difference_kwh",
            "matches_final_cumulative",
        ],
    )

    summary = {
        "input_dir": str(input_dir),
        "output_dir": str(output_dir),
        "files_found": len(files),
        "files_processed": len(metadata_rows),
        "cumulative_rows": len(cumulative_rows),
        "interval_rows": len(interval_rows),
        "validation_events": len(event_rows),
        "daily_validation_rows": len(validation_rows),
        "outputs": [
            "solax_report_metadata.csv",
            "solax_cumulative_rows.csv",
            "solax_interval_energy.csv",
            "solax_validation_events.csv",
            "solax_daily_validation_summary.csv",
        ],
    }
    (output_dir / "solax_processing_summary.json").write_text(
        json.dumps(summary, indent=2), encoding="utf-8"
    )
    return summary


def main() -> int:
    parser = argparse.ArgumentParser(description="Process SolaX Plant Report XLSX files.")
    parser.add_argument("--input-dir", type=Path, default=DEFAULT_INPUT_DIR)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    args = parser.parse_args()

    summary = process_reports(args.input_dir, args.output_dir)
    print(json.dumps(summary, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
